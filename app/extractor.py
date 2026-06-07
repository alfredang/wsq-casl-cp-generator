import re
from pathlib import Path

import openpyxl

from app import config
from app.models import (
    AssessmentMode,
    CourseBackground,
    CourseParticulars,
    CourseSummary,
    ExtractedData,
    InstructionMethod,
    LearningOutcome,
)


def _cell_val(ws, ref: str) -> str:
    """Read a cell value as a stripped string, returning '' if None."""
    val = ws[ref].value
    return str(val).strip() if val is not None else ""


def _extract_particulars(wb) -> CourseParticulars:
    ws = wb[config.SHEET_PARTICULARS]

    # Collect unique skill names from C10:C79
    unique_skills = []
    for row in range(config.CELL_UNIQUE_SKILL_START_ROW, config.UNIQUE_SKILL_MAX_ROW + 1):
        val = ws.cell(row=row, column=3).value  # column C
        if val is None:
            break
        unique_skills.append(str(val).strip())

    return CourseParticulars(
        training_provider=_cell_val(ws, config.CELL_TRAINING_PROVIDER),
        course_title=_cell_val(ws, config.CELL_COURSE_TITLE),
        course_type=_cell_val(ws, config.CELL_COURSE_TYPE),
        about_course=_cell_val(ws, config.CELL_ABOUT_COURSE),
        what_youll_learn=_cell_val(ws, config.CELL_WHAT_YOULL_LEARN),
        unique_skill_names=unique_skills if unique_skills else ["N/A"],
    )


def _extract_background(wb) -> CourseBackground:
    ws = wb[config.SHEET_BACKGROUND]
    return CourseBackground(
        targeted_sectors=_cell_val(ws, config.CELL_TARGETED_SECTORS),
        performance_gaps=_cell_val(ws, config.CELL_PERFORMANCE_GAPS),
    )


def _extract_learning_outcomes(wb) -> list[LearningOutcome]:
    ws = wb[config.SHEET_INSTRUCTIONAL_DESIGN]
    outcomes = []
    row = config.ID_DATA_START_ROW
    while True:
        lo_num = ws[f"{config.ID_COL_LO_NUM}{row}"].value
        if lo_num is None:
            break
        outcomes.append(
            LearningOutcome(
                day=int(ws[f"{config.ID_COL_DAY}{row}"].value or 0),
                duration_minutes=int(ws[f"{config.ID_COL_DURATION}{row}"].value or 0),
                lo_number=str(lo_num).strip(),
                learning_outcome=_cell_val(ws, f"{config.ID_COL_LO_TEXT}{row}"),
                topic=_cell_val(ws, f"{config.ID_COL_TOPIC}{row}").split("\n")[0].strip(),
            )
        )
        row += 1
    return outcomes


def _extract_instruction_methods(wb) -> list[InstructionMethod]:
    ws = wb[config.SHEET_METHODOLOGIES]
    methods = []
    row = config.METH_DATA_START_ROW
    while True:
        method = ws[f"{config.METH_COL_METHOD}{row}"].value
        if method is None:
            break
        methods.append(
            InstructionMethod(
                day=int(ws[f"{config.METH_COL_DAY}{row}"].value or 0),
                method=str(method).strip(),
                duration_minutes=int(ws[f"{config.METH_COL_DURATION}{row}"].value or 0),
                mode_of_training=_cell_val(ws, f"{config.METH_COL_TRAINING_MODE}{row}"),
            )
        )
        row += 1
    return methods


def _extract_assessment_modes(wb) -> list[AssessmentMode]:
    ws = wb[config.SHEET_METHODOLOGIES]
    assessments = []
    row = config.METH_DATA_START_ROW
    while True:
        mode = ws[f"{config.ASSESS_COL_MODE}{row}"].value
        if mode is None:
            break
        assessments.append(
            AssessmentMode(
                day=int(ws[f"{config.ASSESS_COL_DAY}{row}"].value or 0),
                mode=str(mode).strip(),
                duration_minutes=int(ws[f"{config.ASSESS_COL_DURATION}{row}"].value or 0),
                num_assessors=int(ws[f"{config.ASSESS_COL_ASSESSORS}{row}"].value or 0),
                num_candidates=int(ws[f"{config.ASSESS_COL_CANDIDATES}{row}"].value or 0),
            )
        )
        row += 1
    return assessments


def _extract_method_descriptions(wb, name_col: str, desc_col: str) -> dict[str, str]:
    """Read unique method name -> appropriateness elaboration pairs from the
    Methodologies sheet (e.g. G/H for instruction, K/O for assessment)."""
    ws = wb[config.SHEET_METHODOLOGIES]
    descriptions: dict[str, str] = {}
    row = config.METH_DATA_START_ROW
    while True:
        name = ws[f"{name_col}{row}"].value
        if name is None or not str(name).strip():
            break
        desc = ws[f"{desc_col}{row}"].value
        descriptions[str(name).strip()] = str(desc).strip() if desc is not None else ""
        row += 1
    return descriptions


def _extract_summary(wb) -> CourseSummary:
    ws = wb[config.SHEET_SUMMARY]
    return CourseSummary(
        total_course_duration=_cell_val(ws, config.SUMM_TOTAL_COURSE_DURATION),
        total_instructional_duration=_cell_val(ws, config.SUMM_TOTAL_INSTRUCTIONAL),
        total_assessment_duration=_cell_val(ws, config.SUMM_TOTAL_ASSESSMENT),
        mode_of_training=_cell_val(ws, config.SUMM_MODE_OF_TRAINING),
    )


def build_course_topics(data: ExtractedData) -> str:
    """Reconstruct the Course Topics markdown (## Topic N: title / - outcome)
    used on the Course Details page, from an extracted CP's learning outcomes."""
    lines = []
    for i, lo in enumerate(data.learning_outcomes, start=1):
        # Topic cell often looks like "T1: <title>" — drop the T# prefix.
        topic = re.sub(r"^\s*T\d+\s*:\s*", "", lo.topic.strip()).strip()
        lines.append(f"## Topic {i}: {topic}")
        if lo.learning_outcome.strip():
            lines.append(f"- {lo.learning_outcome.strip()}")
        lines.append("")
    return "\n".join(lines).strip()


def build_course_outline(data: ExtractedData) -> str:
    """Build a course outline text (same 3-section format as the AI generator)
    from data extracted out of an existing CP Excel file."""
    lines = ["(1) The list of topics covered in this course"]
    for i, lo in enumerate(data.learning_outcomes, start=1):
        desc = lo.learning_outcome.strip()
        # Topic cell often looks like "T1: <title>" — drop the T# prefix.
        topic = re.sub(r"^\s*T\d+\s*:\s*", "", lo.topic.strip()).strip()
        if desc:
            lines.append(f"T{i}: {topic} - {desc}")
        else:
            lines.append(f"T{i}: {topic}")

    # Section (2): unique instructional methods, preserving order
    lines.append("")
    lines.append("(2) Instructional methods")
    seen = set()
    for im in data.instruction_methods:
        method = im.method.strip()
        if not method or method.lower() in seen:
            continue
        seen.add(method.lower())
        if im.mode_of_training.strip():
            lines.append(f"{method} - {im.mode_of_training.strip()}")
        else:
            lines.append(method)

    # Section (3): duration per topic
    lines.append("")
    lines.append("(3) Duration for each topic")
    for i, lo in enumerate(data.learning_outcomes, start=1):
        lines.append(f"Topic {i}: {lo.duration_minutes}mins")

    return "\n".join(lines)


def extract_data(file_path: Path) -> ExtractedData:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        return ExtractedData(
            particulars=_extract_particulars(wb),
            background=_extract_background(wb),
            learning_outcomes=_extract_learning_outcomes(wb),
            instruction_methods=_extract_instruction_methods(wb),
            assessment_modes=_extract_assessment_modes(wb),
            summary=_extract_summary(wb),
            instruction_method_descriptions=_extract_method_descriptions(
                wb, config.METH_COL_IM_NAME, config.METH_COL_IM_DESC
            ),
            assessment_method_descriptions=_extract_method_descriptions(
                wb, config.METH_COL_AM_NAME, config.METH_COL_AM_DESC
            ),
        )
    finally:
        wb.close()
